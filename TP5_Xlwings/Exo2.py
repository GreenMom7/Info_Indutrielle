#-*- coding:Latin-1 -*-
import xlwings as xw
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import names
from random import *
from openpyxl import Workbook

## Exportation des donn�es dans Excel

# Fonction pour ajouter les notes pour chaque mati�re
def ajouter_notes(feulle_etudiant, matieres, ligne_debut):
    for col_num, matiere in enumerate(matieres, 1):
        feuille_etudiant.cell(row=ligne_debut, column=col_num, value=matiere)

base_donnees = []
for i in range(15):
    etudiant = {"nom": names.get_last_name(),"prenom" : names.get_first_name(), "age" : randint(17,20)}
    base_donnees.append(etudiant)
    
# # Affichage de la base de donn�es
# print("Base de donn�es des �tudiants:")
# for i in range (15):
#     print(f"�tudiant {i}: {base_donnees[i]}")
    

# Cr�ation du classeur Excel et de la feuille "Etudiants"
classeur = Workbook()
feuille_etudiants = classeur.create_sheet("Etudiants", 0)

# Ajout des titres des colonnes
titres = ["nom", "prenom", "age"]
for col_num, titre in enumerate(titres, 1):
    feuille_etudiants.cell(row=1, column=col_num, value=titre)

# Remplissage automatique de la liste des �tudiants
for indice, etudiant in enumerate(base_donnees, 2):  # On commence � la ligne 2 pour les donn�es
    feuille_etudiants['A' + str(indice)].value = etudiant['nom']
    feuille_etudiants['B' + str(indice)].value = etudiant['prenom']
    feuille_etudiants['C' + str(indice)].value = etudiant['age']

# Remplissage des donn�es pour chaque �tudiant
for etudiant in base_donnees:
    nom_prenom = f"{etudiant['nom']} {etudiant['prenom']}"
    
    # Cr�ation d'une feuille pour chaque �tudiant
    feuille_etudiant = classeur.create_sheet(title=nom_prenom)
    
    # Ajout des mati�res en d�but de feuille
    matieres = ["math", "physique", "anglais", "info"]
    ajouter_notes(feuille_etudiant, matieres, ligne_debut=1)
    
    # Ajout des notes pour chaque mati�re (valeurs al�atoires entre 0 et 20)
    for ligne in range(2, 6):
        for col in range(1, 2):
            feuille_etudiant.cell(row=ligne, column=col, value=round(randint(0,20), 2))

# Sauvegarde du classeur Excel
classeur.save("notes_etudiants.xlsx")